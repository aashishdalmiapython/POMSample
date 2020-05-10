import openpyxl

def dataGen():
    #i = [["user1", "user11"],["user2","user22"],["user3","user33"]]
    #return i
    wBook = openpyxl.load_workbook("C:\\FileRead\\openpyxl.xlsx")
    wSheet = wBook["data"]
    row = wSheet.max_row
    cli=[]
    li=[]
    for i in range(1,row+1):
        li=[]
        un = wSheet.cell(i,1).value
        up = wSheet.cell(i,2).value
        li.insert(0,un)
        li.insert(1,up)
        cli.insert(i-1,li)
    return cli